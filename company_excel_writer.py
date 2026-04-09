from __future__ import annotations

import tempfile
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
AI_FILL      = PatternFill("solid", fgColor="FFF2CC")
P1_FILL      = PatternFill("solid", fgColor="FFD6D6")
P2_FILL      = PatternFill("solid", fgColor="FFF3CD")
P3_FILL      = PatternFill("solid", fgColor="E8F5E9")
ISSUES_FILL  = PatternFill("solid", fgColor="FFE8CC")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)

_THIN       = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

PAGE_TYPE_LABELS = {
    "usługa":     "🔧 Usługa",
    "o-nas":      "👥 O nas",
    "kontakt":    "📞 Kontakt",
    "realizacje": "🏆 Realizacje",
    "blog":       "📝 Blog",
    "inne":       "📄 Inne",
}

PRIORITY_LABELS = {1: "🔴 Pilne", 2: "🟡 Ważne", 3: "🟢 Opcjonalne"}
PRIORITY_FILLS  = {1: P1_FILL, 2: P2_FILL, 3: P3_FILL}


def _s(cell, font=None, fill=None, align=None):
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    cell.border = THIN_BORDER


def _banner(ws, text: str, span: str):
    ws.merge_cells(span)
    c = ws[span.split(":")[0]]
    c.value = text
    _s(c, font=HEADER_FONT, fill=HEADER_FILL,
       align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28


def _header_row(ws, row: int, cols: list[tuple[str, int]]):
    for ci, (label, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=label)
        _s(c, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
           align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 32


# ── Sheet 1: Audyt strony firmowej ────────────────────────────────────────────

def _sheet_audit(wb: Workbook, pages: list[dict], ai_result: dict):
    ws = wb.active
    ws.title = "Audyt strony firmowej"

    cols = [
        ("Typ strony",       16),
        ("URL",              38),
        ("Obecny H1",        28),
        ("Rekomendowany H1", 30),
        ("Obecny title",     28),
        ("Rek. title",       30),
        ("Obecny opis",      36),
        ("Rek. opis",        38),
        ("Fraza docelowa",   24),
        ("Problemy",         30),
        ("Priorytet",        16),
    ]
    N = len(cols)
    _banner(ws, "AUDYT SEO — STRONA FIRMOWA", f"A1:{get_column_letter(N)}1")
    _header_row(ws, 2, cols)

    ai_pages_map = {p.get("url", ""): p for p in ai_result.get("pages", [])}
    wrap_top = Alignment(wrap_text=True, vertical="top")

    AI_COLS = {4, 6, 8, 9}  # 1-based columns that get the AI yellow fill

    for ri, page in enumerate(pages, 3):
        url     = page["url"]
        ai      = ai_pages_map.get(url, {})
        prio    = ai.get("priority", 3)
        row_fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL

        page_type_raw   = ai.get("page_type", "inne")
        page_type_label = PAGE_TYPE_LABELS.get(page_type_raw, f"📄 {page_type_raw}")
        issues_text     = "; ".join(ai.get("issues", []))

        row_data = [
            page_type_label,
            url,
            page["h1"],
            ai.get("recommended_h1", ""),
            page["meta_title"],
            ai.get("recommended_title", ""),
            page["meta_description"],
            ai.get("recommended_description", ""),
            ai.get("target_phrase", ""),
            issues_text,
            PRIORITY_LABELS.get(prio, str(prio)),
        ]

        for ci, val in enumerate(row_data, 1):
            if ci == 11:
                cell_fill = PRIORITY_FILLS.get(prio, row_fill)
            elif ci == 10:
                cell_fill = ISSUES_FILL if issues_text else row_fill
            elif ci in AI_COLS:
                cell_fill = AI_FILL
            else:
                cell_fill = row_fill

            cell = ws.cell(row=ri, column=ci, value=val)
            _s(cell, font=BODY_FONT, fill=cell_fill, align=wrap_top)

        ws.row_dimensions[ri].height = 70

    ws.freeze_panes = "A3"


# ── Sheet 2: Quick wins ───────────────────────────────────────────────────────

def _sheet_quick_wins(wb: Workbook, ai_result: dict):
    ws = wb.create_sheet("Quick wins")

    _banner(ws, "QUICK WINS — DO ZROBIENIA W PIERWSZEJ KOLEJNOŚCI", "A1:C1")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30

    summary  = ai_result.get("summary", {})
    wrap_top = Alignment(wrap_text=True, vertical="top")

    row = 2

    def _section_header(title: str):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=title)
        _s(c, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
           align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[row].height = 24
        row += 1

    def _numbered_row(i: int, text: str, fill: PatternFill):
        nonlocal row
        num_cell = ws.cell(row=row, column=1, value=i)
        num_cell.font      = BOLD_FONT
        num_cell.border    = THIN_BORDER
        num_cell.fill      = fill
        num_cell.alignment = Alignment(horizontal="center", vertical="top")

        txt_cell = ws.cell(row=row, column=2, value=text)
        _s(txt_cell, font=BODY_FONT, fill=fill, align=wrap_top)
        ws.merge_cells(f"B{row}:C{row}")
        ws.row_dimensions[row].height = 36
        row += 1

    # Quick wins
    _section_header("⚡ Quick Wins — zrób od razu")
    for i, item in enumerate(summary.get("quick_wins", []), 1):
        _numbered_row(i, item, WHITE_FILL if i % 2 else ALT_FILL)

    row += 1  # blank

    # Biggest issues
    _section_header("🔍 Największe problemy SEO całej strony")
    for i, item in enumerate(summary.get("biggest_issues", []), 1):
        _numbered_row(i, item, P1_FILL if i % 2 else WHITE_FILL)

    row += 1  # blank

    # Priority 1 pages
    p1_pages = [p for p in ai_result.get("pages", []) if p.get("priority") == 1]
    if p1_pages:
        _section_header("🔴 Strony z priorytetem 1 (pilne działania)")
        for p in p1_pages:
            icon_cell = ws.cell(row=row, column=1, value="🔴")
            icon_cell.font      = BOLD_FONT
            icon_cell.border    = THIN_BORDER
            icon_cell.fill      = P1_FILL
            icon_cell.alignment = Alignment(horizontal="center", vertical="top")

            url_cell = ws.cell(row=row, column=2, value=p.get("url", ""))
            _s(url_cell, font=BODY_FONT, fill=P1_FILL, align=wrap_top)

            rsn_cell = ws.cell(row=row, column=3, value=p.get("reasoning", ""))
            _s(rsn_cell, font=BODY_FONT, fill=P1_FILL, align=wrap_top)

            ws.row_dimensions[row].height = 40
            row += 1


# ── Public API ────────────────────────────────────────────────────────────────

def create_company_excel(pages: list[dict], ai_result: dict) -> str:
    """Write audit results to a temp .xlsx file and return its path."""
    wb = Workbook()
    _sheet_audit(wb, pages, ai_result)
    _sheet_quick_wins(wb, ai_result)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def make_company_filename(site_url: str) -> str:
    domain = site_url.replace("https://", "").replace("http://", "").split("/")[0]
    return f"audyt_firmowy_{domain}_{date.today().isoformat()}.xlsx"
