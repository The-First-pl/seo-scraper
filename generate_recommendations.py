from __future__ import annotations

import csv
import os
import time
from collections import defaultdict

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Konfiguracja ──────────────────────────────────────────────────────────────

INPUT_CSV = "categories_with_gsc.csv"
OUTPUT_XLSX = "seo_audit_kategorii.xlsx"

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# ── Wczytanie CSV i grupowanie po URL ─────────────────────────────────────────

def load_categories(csv_path: str) -> dict[str, dict]:
    """
    Zwraca słownik: url → {meta, queries: [{query, clicks, impressions, position}]}
    """
    categories = defaultdict(lambda: {
        "url": "",
        "h1": "",
        "meta_title": "",
        "meta_description": "",
        "queries": [],
    })

    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            url = row["url"].strip()
            cat = categories[url]
            cat["url"] = url
            cat["h1"] = row["h1"]
            cat["meta_title"] = row["meta_title"]
            cat["meta_description"] = row["meta_description"]
            if row["gsc_query"]:
                cat["queries"].append({
                    "query": row["gsc_query"],
                    "clicks": int(row["gsc_clicks"] or 0),
                    "impressions": int(row["gsc_impressions"] or 0),
                    "position": float(row["gsc_position"] or 0),
                })

    # Sortuj frazy po kliknięciach malejąco
    for cat in categories.values():
        cat["queries"].sort(key=lambda x: x["clicks"], reverse=True)

    return dict(categories)

# ── Prompt do Claude ──────────────────────────────────────────────────────────

def build_prompt(cat: dict) -> str:
    top_queries = cat["queries"][:20]  # max 20 fraz do kontekstu

    queries_text = ""
    if top_queries:
        lines = []
        for q in top_queries:
            lines.append(
                f"  - \"{q['query']}\" | kliknięcia: {q['clicks']} | "
                f"wyświetlenia: {q['impressions']} | pozycja: {q['position']}"
            )
        queries_text = "\n".join(lines)
    else:
        queries_text = "  (brak danych)"

    return f"""Jesteś ekspertem SEO dla polskich sklepów internetowych. Przeanalizuj poniższą kategorię sklepu z kawą i podaj konkretne rekomendacje.

## Dane kategorii

- **URL:** {cat['url']}
- **Obecny H1:** {cat['h1'] or '(brak)'}
- **Obecny meta title:** {cat['meta_title'] or '(brak)'}
- **Obecny meta description:** {cat['meta_description'] or '(brak)'}

## Frazy z Google Search Console (ostatnie 90 dni)

{queries_text}

## Zadanie

Na podstawie powyższych danych przygotuj rekomendacje SEO. Odpowiedz WYŁĄCZNIE w formacie JSON (bez markdown, bez komentarzy):

{{
  "h1_rekomendacja": "nowy tekst H1 (max 60 znaków, zawierający główną frazę kluczową)",
  "h1_uzasadnienie": "krótkie uzasadnienie zmiany (1-2 zdania)",
  "meta_title_rekomendacja": "nowy meta title (max 60 znaków)",
  "meta_description_rekomendacja": "nowy meta description (max 155 znaków, z CTA)",
  "frazy_h2": ["fraza 1 do H2", "fraza 2 do H2", "fraza 3 do H2", "fraza 4 do H2", "fraza 5 do H2"],
  "propozycje_podkategorii": ["podkategoria 1", "podkategoria 2", "podkategoria 3"],
  "glowna_fraza": "najważniejsza fraza kluczowa dla tej kategorii",
  "ocena_obecnego_seo": "krótka ocena obecnego stanu SEO (1-2 zdania)"
}}"""


def get_claude_recommendations(client: anthropic.Anthropic, cat: dict) -> dict:
    try:
        message = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": build_prompt(cat)}],
        )
        raw = message.content[0].text.strip()

        # Wyciągnij JSON jeśli jest otoczony markdown
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]

        import json
        return json.loads(raw)
    except Exception as e:
        print(f"  BŁĄD Claude dla {cat['url']}: {e}")
        return {
            "h1_rekomendacja": "",
            "h1_uzasadnienie": f"Błąd API: {e}",
            "meta_title_rekomendacja": "",
            "meta_description_rekomendacja": "",
            "frazy_h2": [],
            "propozycje_podkategorii": [],
            "glowna_fraza": "",
            "ocena_obecnego_seo": "",
        }

# ── Excel ─────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # ciemny granat
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # niebieski
GSC_FILL      = PatternFill("solid", fgColor="E2EFDA")   # jasna zieleń
AI_FILL       = PatternFill("solid", fgColor="FFF2CC")   # jasny żółty
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="F5F5F5")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_cell(cell, font=None, fill=None, alignment=None, border=True):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER


def create_summary_sheet(wb: Workbook, categories: dict, recommendations: dict):
    ws = wb.active
    ws.title = "Przegląd kategorii"

    # Nagłówek główny
    ws.merge_cells("A1:N1")
    cell = ws["A1"]
    cell.value = "AUDYT SEO KATEGORII — kawa-amann.pl"
    style_cell(cell, font=Font(name="Calibri", bold=True, color="FFFFFF", size=14),
               fill=HEADER_FILL, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # Nagłówki kolumn
    headers = [
        ("URL kategorii", 40),
        ("Obecny H1", 30),
        ("Rekomendowany H1", 30),
        ("Uzasadnienie H1", 35),
        ("Obecny meta title", 30),
        ("Rekomendowany meta title", 30),
        ("Obecny meta desc", 40),
        ("Rekomendowany meta desc", 40),
        ("Główna fraza", 25),
        ("Ocena SEO", 35),
        ("Frazy do H2", 40),
        ("Propozycje podkategorii", 35),
        ("Liczba fraz GSC", 15),
        ("Keyword Planner (placeholder)", 25),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        style_cell(cell, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 35

    # Dane
    for row_idx, (url, cat) in enumerate(sorted(categories.items()), 3):
        rec = recommendations.get(url, {})
        fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL

        frazy_h2 = "\n".join(f"• {f}" for f in rec.get("frazy_h2", []))
        podkategorie = "\n".join(f"• {p}" for p in rec.get("propozycje_podkategorii", []))

        row_data = [
            url,
            cat["h1"],
            rec.get("h1_rekomendacja", ""),
            rec.get("h1_uzasadnienie", ""),
            cat["meta_title"],
            rec.get("meta_title_rekomendacja", ""),
            cat["meta_description"],
            rec.get("meta_description_rekomendacja", ""),
            rec.get("glowna_fraza", ""),
            rec.get("ocena_obecnego_seo", ""),
            frazy_h2,
            podkategorie,
            len(cat["queries"]),
            "",  # Keyword Planner placeholder
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Kolumny AI — żółte tło
            if col_idx in (3, 4, 6, 8, 9, 10, 11, 12):
                style_cell(cell, font=BODY_FONT, fill=AI_FILL,
                           alignment=Alignment(wrap_text=True, vertical="top"))
            else:
                style_cell(cell, font=BODY_FONT, fill=fill,
                           alignment=Alignment(wrap_text=True, vertical="top"))

        ws.row_dimensions[row_idx].height = 80

    # Zablokuj nagłówki
    ws.freeze_panes = "A3"


def create_gsc_sheet(wb: Workbook, categories: dict):
    ws = wb.create_sheet("Dane GSC")

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "DANE GOOGLE SEARCH CONSOLE — frazy per kategoria"
    style_cell(cell, font=Font(name="Calibri", bold=True, color="FFFFFF", size=13),
               fill=HEADER_FILL, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    headers = [
        ("URL kategorii", 45),
        ("H1", 30),
        ("Fraza", 35),
        ("Kliknięcia", 14),
        ("Wyświetlenia", 14),
        ("Pozycja śr.", 13),
        ("Keyword Planner (placeholder)", 28),
    ]
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        style_cell(cell, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    row_idx = 3
    for url, cat in sorted(categories.items()):
        queries = cat["queries"]
        if not queries:
            cell_url = ws.cell(row=row_idx, column=1, value=url)
            style_cell(cell_url, font=BODY_FONT, fill=WHITE_FILL)
            ws.cell(row=row_idx, column=2, value=cat["h1"])
            ws.cell(row=row_idx, column=3, value="(brak danych GSC)")
            row_idx += 1
            continue

        for q_idx, q in enumerate(queries):
            fill = GSC_FILL if q_idx % 2 == 0 else WHITE_FILL
            row_data = [
                url if q_idx == 0 else "",
                cat["h1"] if q_idx == 0 else "",
                q["query"],
                q["clicks"],
                q["impressions"],
                q["position"],
                "",  # Keyword Planner
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                align = Alignment(vertical="top", wrap_text=(col_idx <= 2))
                style_cell(cell, font=BODY_FONT, fill=fill, alignment=align)
            row_idx += 1

    ws.freeze_panes = "A3"


def create_ai_details_sheet(wb: Workbook, categories: dict, recommendations: dict):
    ws = wb.create_sheet("Rekomendacje AI (szczegóły)")

    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = "REKOMENDACJE AI — szczegóły per kategoria"
    style_cell(cell, font=Font(name="Calibri", bold=True, color="FFFFFF", size=13),
               fill=HEADER_FILL, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 70

    row = 2
    for url, cat in sorted(categories.items()):
        rec = recommendations.get(url, {})

        # Nagłówek kategorii
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=url)
        style_cell(cell, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
                   alignment=Alignment(wrap_text=True, vertical="center"))
        ws.row_dimensions[row].height = 22
        row += 1

        fields = [
            ("Obecny H1", cat["h1"]),
            ("→ Rekomendowany H1", rec.get("h1_rekomendacja", "")),
            ("Uzasadnienie H1", rec.get("h1_uzasadnienie", "")),
            ("Obecny meta title", cat["meta_title"]),
            ("→ Rekomendowany meta title", rec.get("meta_title_rekomendacja", "")),
            ("Obecny meta description", cat["meta_description"]),
            ("→ Rekomendowany meta desc", rec.get("meta_description_rekomendacja", "")),
            ("Główna fraza kluczowa", rec.get("glowna_fraza", "")),
            ("Ocena obecnego SEO", rec.get("ocena_obecnego_seo", "")),
        ]

        for label, value in fields:
            is_rec = label.startswith("→")
            fill = AI_FILL if is_rec else WHITE_FILL
            font = BOLD_FONT if is_rec else BODY_FONT

            ws.cell(row=row, column=1)  # pusta
            cell_label = ws.cell(row=row, column=2, value=label)
            cell_value = ws.cell(row=row, column=3, value=value)
            style_cell(cell_label, font=font, fill=fill,
                       alignment=Alignment(vertical="top", wrap_text=True))
            style_cell(cell_value, font=font, fill=fill,
                       alignment=Alignment(vertical="top", wrap_text=True))
            ws.row_dimensions[row].height = 40
            row += 1

        # Frazy do H2
        cell_label = ws.cell(row=row, column=2, value="Frazy do sekcji H2")
        style_cell(cell_label, font=BOLD_FONT, fill=AI_FILL,
                   alignment=Alignment(vertical="top", wrap_text=True))
        frazy = rec.get("frazy_h2", [])
        cell_val = ws.cell(row=row, column=3, value="\n".join(f"• {f}" for f in frazy))
        style_cell(cell_val, font=BODY_FONT, fill=AI_FILL,
                   alignment=Alignment(vertical="top", wrap_text=True))
        ws.row_dimensions[row].height = max(40, 15 * len(frazy))
        row += 1

        # Propozycje podkategorii
        cell_label = ws.cell(row=row, column=2, value="Propozycje podkategorii")
        style_cell(cell_label, font=BOLD_FONT, fill=AI_FILL,
                   alignment=Alignment(vertical="top", wrap_text=True))
        podkat = rec.get("propozycje_podkategorii", [])
        cell_val = ws.cell(row=row, column=3, value="\n".join(f"• {p}" for p in podkat))
        style_cell(cell_val, font=BODY_FONT, fill=AI_FILL,
                   alignment=Alignment(vertical="top", wrap_text=True))
        ws.row_dimensions[row].height = max(40, 15 * len(podkat))
        row += 1

        # Odstęp między kategoriami
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not ANTHROPIC_API_KEY:
        print("BŁĄD: Ustaw zmienną środowiskową ANTHROPIC_API_KEY")
        print("  export ANTHROPIC_API_KEY='sk-ant-...'")
        return

    print("=== KROK 1: Wczytywanie CSV ===")
    categories = load_categories(INPUT_CSV)
    print(f"Wczytano {len(categories)} kategorii, łącznie {sum(len(c['queries']) for c in categories.values())} fraz GSC\n")

    print("=== KROK 2: Generowanie rekomendacji przez Claude API ===")
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    recommendations = {}

    for i, (url, cat) in enumerate(sorted(categories.items()), 1):
        print(f"[{i}/{len(categories)}] {url.split('/')[-1]}")
        rec = get_claude_recommendations(client, cat)
        recommendations[url] = rec
        print(f"  H1: {rec.get('h1_rekomendacja', '?')[:60]}")
        print(f"  Frazy H2: {len(rec.get('frazy_h2', []))} | Podkategorie: {len(rec.get('propozycje_podkategorii', []))}")
        time.sleep(0.5)  # rate limit

    print(f"\n=== KROK 3: Tworzenie pliku Excel: {OUTPUT_XLSX} ===")
    wb = Workbook()

    create_summary_sheet(wb, categories, recommendations)
    create_gsc_sheet(wb, categories)
    create_ai_details_sheet(wb, categories, recommendations)

    wb.save(OUTPUT_XLSX)
    print(f"\nGotowe! Zapisano do: {OUTPUT_XLSX}")
    print(f"Arkusze: 'Przegląd kategorii' | 'Dane GSC' | 'Rekomendacje AI (szczegóły)'")


if __name__ == "__main__":
    main()
