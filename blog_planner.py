from __future__ import annotations

import json
import re
import tempfile

import anthropic
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from scraper import _get, scrape_page_meta, discover_categories

# ── Scraping site context ─────────────────────────────────────────────────────

def scrape_site_context(url: str) -> dict:
    """
    Collect basic SEO signals from the given URL to enrich the Claude prompt.
    Returns a dict with title, description, and a sample of category names.
    """
    meta = scrape_page_meta(url)
    try:
        categories = discover_categories(url)
        cat_names = [u.rstrip("/").split("/")[-1].replace("-", " ") for u in categories[:15]]
    except Exception:
        cat_names = []

    return {
        "title": meta.get("meta_title", ""),
        "description": meta.get("meta_description", ""),
        "categories": cat_names,
    }


# ── Claude prompt & parsing ───────────────────────────────────────────────────

_PROMPT_TEMPLATE = """Jesteś ekspertem SEO i content marketingu. Na podstawie poniższych danych zaplanuj blog.

Tematyka: {tematyka}
Język: {jezyk}
Liczba klastrów tematycznych: {liczba_klastrow}
{site_section}
Wygeneruj plan bloga w formacie JSON:
{{
  "clusters": [
    {{
      "name": "Nazwa kontenera tematycznego",
      "pillar": {{
        "title": "Tytuł wpisu głównego (pillar)",
        "main_phrase": "fraza główna",
        "intent": "informacyjna|zakupowa|nawigacyjna|posprzedażowa",
        "priority": 1
      }},
      "supporting": [
        {{
          "title": "Tytuł wpisu wspierającego",
          "main_phrase": "fraza wspierająca",
          "intent": "informacyjna|zakupowa|nawigacyjna|posprzedażowa",
          "priority": 2
        }}
      ]
    }}
  ]
}}

Zasady:
- Liczba klastrów musi wynosić dokładnie {liczba_klastrow}
- Każdy klaster ma dokładnie 1 pillar + od 4 do 8 supporting
- Frazy mają być realistyczne, w języku: {jezyk}
- Intencja musi być dopasowana do frazy (informacyjna dla "jak", zakupowa dla "kup/sklep/cena", nawigacyjna dla brandowych, posprzedażowa dla "opinie/recenzja")
- Pillar zawsze priority: 1, supporting: 2–4 według ważności
- Zwróć TYLKO JSON, bez żadnego tekstu przed ani po"""

_RETRY_SUFFIX = "\n\nUWAGA: Poprzednia odpowiedź zawierała niepoprawny JSON. Zwróć TYLKO poprawny JSON, nic więcej."


def _build_prompt(tematyka: str, jezyk: str, liczba: int, site_ctx: dict | None) -> str:
    if site_ctx:
        parts = []
        if site_ctx.get("title"):
            parts.append(f"Tytuł strony: {site_ctx['title']}")
        if site_ctx.get("description"):
            parts.append(f"Meta description: {site_ctx['description']}")
        if site_ctx.get("categories"):
            parts.append("Istniejące kategorie/podstrony: " + ", ".join(site_ctx["categories"]))
        site_section = "Dane strony:\n" + "\n".join(parts) + "\n\n" if parts else ""
    else:
        site_section = ""

    return _PROMPT_TEMPLATE.format(
        tematyka=tematyka,
        jezyk=jezyk,
        liczba_klastrow=liczba,
        site_section=site_section,
    )


def _parse_json(text: str) -> dict:
    """Extract and parse JSON from Claude's response, handling markdown fences."""
    text = text.strip()
    # Strip ```json ... ``` fences
    fence = re.search(r"```(?:json)?\s*([\s\S]+?)```", text)
    if fence:
        text = fence.group(1).strip()
    return json.loads(text)


def generate_blog_plan(
    client: anthropic.Anthropic,
    tematyka: str,
    jezyk: str,
    liczba: int,
    site_ctx: dict | None,
) -> dict:
    """
    Calls Claude to generate the blog plan. Retries once with a stricter prompt
    if the first response cannot be parsed as JSON.
    """
    prompt = _build_prompt(tematyka, jezyk, liczba, site_ctx)

    for attempt in range(2):
        msg = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt + (_RETRY_SUFFIX if attempt else "")}],
        )
        raw = msg.content[0].text
        try:
            return _parse_json(raw)
        except (json.JSONDecodeError, ValueError):
            if attempt == 1:
                raise ValueError(
                    f"Claude zwrócił niepoprawny JSON po 2 próbach.\n\nOdpowiedź:\n{raw[:500]}"
                )

    raise RuntimeError("unreachable")


# ── Excel export ──────────────────────────────────────────────────────────────

_HDR_FILL    = PatternFill("solid", fgColor="404040")
_PILLAR_FILL = PatternFill("solid", fgColor="FFF9C4")   # light yellow
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_CLUSTER_FILLS = [
    PatternFill("solid", fgColor="E3F2FD"),  # blue-50
    PatternFill("solid", fgColor="E8F5E9"),  # green-50
    PatternFill("solid", fgColor="F3E5F5"),  # purple-50
    PatternFill("solid", fgColor="FFF3E0"),  # orange-50
    PatternFill("solid", fgColor="FCE4EC"),  # pink-50
]

_HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT  = Font(name="Calibri", size=10)
_BOLD_FONT  = Font(name="Calibri", bold=True, size=10)

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

INTENT_LABELS = {
    "informacyjna":    "ℹ️ Informacyjna",
    "zakupowa":        "🛒 Zakupowa",
    "nawigacyjna":     "🧭 Nawigacyjna",
    "posprzedażowa":   "⭐ Posprzedażowa",
}

COLUMNS = [
    ("Klaster",       28),
    ("Typ",           16),
    ("Tytuł wpisu",   52),
    ("Fraza główna",  34),
    ("Intencja",      20),
    ("Priorytet",     11),
    ("Status",        18),
]


def _cell(ws, row: int, col: int, value, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = font  or _BODY_FONT
    c.fill   = fill  or _WHITE_FILL
    c.border = _BORDER
    c.alignment = align or Alignment(vertical="top", wrap_text=True)
    return c


def export_excel(plan: dict, tematyka: str) -> str:
    """Serialize the blog plan to an .xlsx temp file. Returns the file path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan bloga"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.border    = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 2
    clusters = plan.get("clusters", [])

    for cl_idx, cluster in enumerate(clusters):
        cluster_fill = _CLUSTER_FILLS[cl_idx % len(_CLUSTER_FILLS)]
        cluster_name = cluster.get("name", f"Klaster {cl_idx + 1}")
        pillar       = cluster.get("pillar", {})
        supporting   = cluster.get("supporting", [])
        total_rows   = 1 + len(supporting)
        first_row    = row

        # Pillar row
        _cell(ws, row, 1, cluster_name, font=_BOLD_FONT, fill=cluster_fill,
              align=Alignment(vertical="center", wrap_text=True, horizontal="center"))
        _cell(ws, row, 2, "🟢 Pillar",    font=_BOLD_FONT, fill=_PILLAR_FILL)
        _cell(ws, row, 3, pillar.get("title", ""),        font=_BOLD_FONT, fill=_PILLAR_FILL)
        _cell(ws, row, 4, pillar.get("main_phrase", ""),  font=_BOLD_FONT, fill=_PILLAR_FILL)
        _cell(ws, row, 5,
              INTENT_LABELS.get(pillar.get("intent", ""), pillar.get("intent", "")),
              fill=_PILLAR_FILL)
        _cell(ws, row, 6, pillar.get("priority", 1),      fill=_PILLAR_FILL,
              align=Alignment(horizontal="center", vertical="top"))
        _cell(ws, row, 7, "", fill=_PILLAR_FILL)
        ws.row_dimensions[row].height = 36
        row += 1

        # Supporting rows
        for art in supporting:
            _cell(ws, row, 1, "", fill=cluster_fill)   # blank — visually grouped by cluster color
            _cell(ws, row, 2, "🔵 Supporting")
            _cell(ws, row, 3, art.get("title", ""))
            _cell(ws, row, 4, art.get("main_phrase", ""))
            _cell(ws, row, 5, INTENT_LABELS.get(art.get("intent", ""), art.get("intent", "")))
            _cell(ws, row, 6, art.get("priority", 2),
                  align=Alignment(horizontal="center", vertical="top"))
            _cell(ws, row, 7, "")
            ws.row_dimensions[row].height = 30
            row += 1

        # Merge cluster name cell vertically
        if total_rows > 1:
            ws.merge_cells(
                start_row=first_row, start_column=1,
                end_row=first_row + total_rows - 1, end_column=1,
            )

    # ── Summary tab ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Podsumowanie")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    summary_hdr = [("Klaster", "Łącznie wpisów")]
    for col_idx, label in enumerate(summary_hdr[0], 1):
        c = ws2.cell(row=1, column=col_idx, value=label)
        c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"

    total_articles = 0
    for r_idx, cluster in enumerate(clusters, 2):
        count = 1 + len(cluster.get("supporting", []))
        total_articles += count
        c1 = ws2.cell(row=r_idx, column=1, value=cluster.get("name", ""))
        c1.font = _BODY_FONT; c1.border = _BORDER
        c1.alignment = Alignment(wrap_text=True, vertical="top")
        c2 = ws2.cell(row=r_idx, column=2, value=count)
        c2.font = _BODY_FONT; c2.border = _BORDER
        c2.alignment = Alignment(horizontal="center")

    # Total row
    tr = len(clusters) + 2
    c1 = ws2.cell(row=tr, column=1, value="RAZEM")
    c1.font = _BOLD_FONT; c1.border = _BORDER
    c2 = ws2.cell(row=tr, column=2, value=total_articles)
    c2.font = _BOLD_FONT; c2.border = _BORDER
    c2.alignment = Alignment(horizontal="center")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
