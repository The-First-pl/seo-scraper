from __future__ import annotations

import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style ─────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
GSC_FILL     = PatternFill("solid", fgColor="E2EFDA")
AI_FILL      = PatternFill("solid", fgColor="FFF2CC")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)

_THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style(cell, font=None, fill=None, align=None):
    if font:  cell.font   = font
    if fill:  cell.fill   = fill
    if align: cell.alignment = align
    cell.border = THIN_BORDER


def _header_row(ws, row: int, cols: list[tuple[str, int]]):
    for col_idx, (label, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=label)
        _style(c, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 32


def _banner(ws, text: str, span: str):
    ws.merge_cells(span)
    c = ws[span.split(":")[0]]
    c.value = text
    _style(c, font=HEADER_FONT, fill=HEADER_FILL,
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28


# ── Sheet 1: Przegląd ─────────────────────────────────────────────────────────

def _sheet_overview(wb: Workbook, categories: dict, recommendations: dict):
    ws = wb.active
    ws.title = "Przegląd kategorii"

    cols = [
        ("URL kategorii", 42), ("Obecny H1", 28), ("Rekomendowany H1", 30),
        ("Uzasadnienie H1", 34), ("Obecny meta title", 28), ("Rek. meta title", 30),
        ("Obecny meta desc", 38), ("Rek. meta desc", 40), ("Główna fraza", 24),
        ("Ocena SEO", 34), ("Frazy do H2", 38), ("Propozycje podkategorii", 34),
        ("Fraz GSC", 11), ("Keyword Planner", 22),
    ]
    _banner(ws, "AUDYT SEO KATEGORII", f"A1:{get_column_letter(len(cols))}1")
    _header_row(ws, 2, cols)

    AI_COLS = {3, 4, 6, 8, 9, 10, 11, 12}
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for r, (url, cat) in enumerate(sorted(categories.items()), 3):
        rec = recommendations.get(url, {})
        fill = WHITE_FILL if r % 2 == 0 else ALT_FILL

        row_data = [
            url,
            cat["h1"],
            rec.get("h1_rekomendacja", ""),
            rec.get("h1_uzasadnienie", ""),
            cat["meta_title"],
            rec.get("meta_title_rekomendacja", ""),
            cat["meta_description"],
            rec.get("meta_description_rekomendacja", ""),
            rec.get("glowna_fraza", ""),
            rec.get("ocena_obecnego_seo", ""),
            "\n".join(f"• {f}" for f in rec.get("frazy_h2", [])),
            "\n".join(f"• {p}" for p in rec.get("propozycje_podkategorii", [])),
            len(cat["queries"]),
            "",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            _style(cell, font=BODY_FONT,
                   fill=AI_FILL if c in AI_COLS else fill,
                   align=wrap_top)
        ws.row_dimensions[r].height = 75

    ws.freeze_panes = "A3"


# ── Sheet 2: Dane GSC ─────────────────────────────────────────────────────────

def _sheet_gsc(wb: Workbook, categories: dict):
    ws = wb.create_sheet("Dane GSC")
    cols = [
        ("URL kategorii", 44), ("H1", 28), ("Fraza", 34),
        ("Kliknięcia", 13), ("Wyświetlenia", 13), ("Pozycja śr.", 12), ("Keyword Planner", 24),
    ]
    _banner(ws, "DANE GOOGLE SEARCH CONSOLE", f"A1:{get_column_letter(len(cols))}1")
    _header_row(ws, 2, cols)

    row = 3
    wrap_top = Alignment(vertical="top", wrap_text=True)
    for url, cat in sorted(categories.items()):
        queries = cat["queries"]
        if not queries:
            for c, val in enumerate([url, cat["h1"], "(brak danych GSC)", "", "", "", ""], 1):
                _style(ws.cell(row=row, column=c, value=val), font=BODY_FONT, fill=WHITE_FILL)
            row += 1
            continue

        for qi, q in enumerate(queries):
            fill = GSC_FILL if qi % 2 == 0 else WHITE_FILL
            for c, val in enumerate([
                url if qi == 0 else "",
                cat["h1"] if qi == 0 else "",
                q["query"], q["clicks"], q["impressions"], q["position"], "",
            ], 1):
                _style(ws.cell(row=row, column=c, value=val),
                       font=BODY_FONT, fill=fill, align=wrap_top)
            row += 1

    ws.freeze_panes = "A3"


# ── Sheet 3: Szczegóły AI ─────────────────────────────────────────────────────

def _sheet_ai(wb: Workbook, categories: dict, recommendations: dict):
    ws = wb.create_sheet("Rekomendacje AI")
    _banner(ws, "REKOMENDACJE AI — szczegóły per kategoria", "A1:C1")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 68

    row = 2
    wrap_top = Alignment(vertical="top", wrap_text=True)

    for url, cat in sorted(categories.items()):
        rec = recommendations.get(url, {})

        # Category header
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=url)
        _style(c, font=SUBHEAD_FONT, fill=SUBHEAD_FILL,
               align=Alignment(wrap_text=True, vertical="center"))
        ws.row_dimensions[row].height = 20
        row += 1

        fields = [
            (False, "Obecny H1",               cat["h1"]),
            (True,  "→ Rekomendowany H1",       rec.get("h1_rekomendacja", "")),
            (True,  "Uzasadnienie",             rec.get("h1_uzasadnienie", "")),
            (False, "Obecny meta title",        cat["meta_title"]),
            (True,  "→ Rek. meta title",        rec.get("meta_title_rekomendacja", "")),
            (False, "Obecny meta description",  cat["meta_description"]),
            (True,  "→ Rek. meta description",  rec.get("meta_description_rekomendacja", "")),
            (True,  "Główna fraza",             rec.get("glowna_fraza", "")),
            (True,  "Ocena SEO",                rec.get("ocena_obecnego_seo", "")),
        ]
        for is_rec, label, value in fields:
            fill = AI_FILL if is_rec else WHITE_FILL
            font = BOLD_FONT if is_rec else BODY_FONT
            _style(ws.cell(row=row, column=2, value=label), font=font, fill=fill, align=wrap_top)
            _style(ws.cell(row=row, column=3, value=value), font=font, fill=fill, align=wrap_top)
            ws.row_dimensions[row].height = 38
            row += 1

        # Frazy H2
        frazy = rec.get("frazy_h2", [])
        _style(ws.cell(row=row, column=2, value="Frazy do sekcji H2"),
               font=BOLD_FONT, fill=AI_FILL, align=wrap_top)
        _style(ws.cell(row=row, column=3, value="\n".join(f"• {f}" for f in frazy)),
               font=BODY_FONT, fill=AI_FILL, align=wrap_top)
        ws.row_dimensions[row].height = max(38, 14 * len(frazy))
        row += 1

        # Podkategorie
        podkat = rec.get("propozycje_podkategorii", [])
        _style(ws.cell(row=row, column=2, value="Propozycje podkategorii"),
               font=BOLD_FONT, fill=AI_FILL, align=wrap_top)
        _style(ws.cell(row=row, column=3, value="\n".join(f"• {p}" for p in podkat)),
               font=BODY_FONT, fill=AI_FILL, align=wrap_top)
        ws.row_dimensions[row].height = max(38, 14 * len(podkat))
        row += 2  # blank row between categories


# ── Public API ────────────────────────────────────────────────────────────────

def create_excel(categories: dict, recommendations: dict) -> str:
    """Creates Excel file and returns the path to a temp file."""
    wb = Workbook()
    _sheet_overview(wb, categories, recommendations)
    _sheet_gsc(wb, categories)
    _sheet_ai(wb, categories, recommendations)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
